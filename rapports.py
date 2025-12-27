from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os

BASE_EXPORT_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "exports")
if not os.path.exists(BASE_EXPORT_DIR): os.makedirs(BASE_EXPORT_DIR)

# Mapping des noms pour le PDF
NOMS_COMPTES = {
    1111: "Capital Social",
    1481: "Emprunts bancaires",
    2111: "Frais préliminaires",
    2340: "Matériel de transport",
    2351: "Mobilier de bureau",
    2352: "Matériel informatique",
    3111: "Marchandises (Stock)",
    3421: "Clients",
    3455: "État TVA Récupérable",
    4411: "Fournisseurs",
    4455: "État TVA Facturée",
    5141: "Banque (DH)",
    5161: "Caisse",
    6111: "Achats marchandises",
    6131: "Location",
    6171: "Salaires",
    7111: "Ventes marchandises"
}

class GenerateurRapport:
    def _creer_dossier_entreprise(self, nom_entreprise):
        nom = nom_entreprise.strip()
        chemin = os.path.join(BASE_EXPORT_DIR, nom)
        if not os.path.exists(chemin): os.makedirs(chemin)
        return chemin

    def exporter_journal_excel(self, data, nom_entreprise):
        try:
            dossier = self._creer_dossier_entreprise(nom_entreprise)
            filepath = os.path.join(dossier, f"Journal_{nom_entreprise}.xlsx")
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Journal"
            ws.append(["Date", "Libellé", "Compte", "Débit", "Crédit", "Référence"])

            header_fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            for row in data:
                ws.append(row)

            # Largeurs
            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['F'].width = 20
            
            wb.save(filepath)
            return filepath
        except Exception as e:
            print(e)
            return None

    def generer_bilan_pdf(self, nom_entreprise, lignes_actif, totaux_actif, lignes_passif, total_passif):
        try:
            dossier = self._creer_dossier_entreprise(nom_entreprise)
            date_jour = datetime.now().strftime("%Y-%m-%d")
            filepath = os.path.join(dossier, f"Bilan_{nom_entreprise}_{date_jour}.pdf")

            pdf = FPDF()
            pdf.set_auto_page_break(auto=True, margin=15)

            # --- PAGE 1 : ACTIF ---
            pdf.add_page()
            pdf.set_font("Arial", 'B', 16)
            pdf.cell(0, 10, txt="BILAN - ACTIF (Page 1)", ln=True, align='C')
            pdf.set_font("Arial", 'I', 10)
            pdf.cell(0, 10, txt=f"Entreprise : {nom_entreprise} | Date : {date_jour}", ln=True, align='C')
            pdf.ln(10)

            # Tableau Actif
            pdf.set_fill_color(220, 220, 220)
            pdf.set_font("Arial", 'B', 10)
            pdf.cell(80, 10, "COMPTE", 1, 0, 'C', True)
            pdf.cell(35, 10, "BRUT", 1, 0, 'C', True)
            pdf.cell(35, 10, "AMORT", 1, 0, 'C', True)
            pdf.cell(35, 10, "NET", 1, 1, 'C', True)

            pdf.set_font("Arial", size=10)
            for l in lignes_actif:
                nom = NOMS_COMPTES.get(l['compte'], f"Cpt {l['compte']}")
                pdf.cell(80, 8, f"{l['compte']} - {nom}", 1)
                pdf.cell(35, 8, f"{l['brut']:.2f}", 1, 0, 'R')
                pdf.cell(35, 8, f"{l['amort']:.2f}", 1, 0, 'R')
                pdf.cell(35, 8, f"{l['net']:.2f}", 1, 1, 'R')

            # Total Actif
            pdf.ln(5)
            pdf.set_font("Arial", 'B', 11)
            pdf.set_fill_color(200, 255, 200)
            pdf.cell(80, 10, "TOTAL ACTIF", 1, 0, 'C', True)
            pdf.cell(35, 10, f"{totaux_actif['brut']:.2f}", 1, 0, 'R', True)
            pdf.cell(35, 10, f"{totaux_actif['amort']:.2f}", 1, 0, 'R', True)
            pdf.cell(35, 10, f"{totaux_actif['net']:.2f}", 1, 1, 'R', True)

            # --- PAGE 2 : PASSIF ---
            pdf.add_page()
            pdf.set_font("Arial", 'B', 16)
            pdf.cell(0, 10, txt="BILAN - PASSIF (Page 2)", ln=True, align='C')
            pdf.ln(10)

            # Tableau Passif
            pdf.set_fill_color(220, 220, 220)
            pdf.set_font("Arial", 'B', 10)
            pdf.cell(130, 10, "COMPTE", 1, 0, 'C', True)
            pdf.cell(55, 10, "NET", 1, 1, 'C', True)

            pdf.set_font("Arial", size=10)
            for l in lignes_passif:
                if l['compte'] == "RÉSULTAT NET":
                    txt = "RÉSULTAT NET DE L'EXERCICE"
                else:
                    nom = NOMS_COMPTES.get(l['compte'], f"Cpt {l['compte']}")
                    txt = f"{l['compte']} - {nom}"
                
                pdf.cell(130, 8, txt, 1)
                pdf.cell(55, 8, f"{l['net']:.2f}", 1, 1, 'R')

            # Total Passif
            pdf.ln(5)
            pdf.set_font("Arial", 'B', 11)
            pdf.set_fill_color(255, 200, 200)
            pdf.cell(130, 10, "TOTAL PASSIF", 1, 0, 'C', True)
            pdf.cell(55, 10, f"{total_passif:.2f}", 1, 1, 'R', True)

            pdf.output(filepath)
            return filepath
        except Exception as e:
            print(f"PDF Error: {e}")
            return None