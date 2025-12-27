from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os

BASE_EXP = os.path.join(os.path.dirname(os.path.dirname(__file__)), "exports")
if not os.path.exists(BASE_EXP): os.makedirs(BASE_EXP)

# --- COULEURS DU THÈME ---
COLOR_THEME = (0, 169, 181)    # Le Bleu/Cyan du Journal
COLOR_LOGO = (243, 156, 18)    # L'Orange du Logo
COLOR_TEXT_HEAD = (255, 255, 255) # Blanc
COLOR_LIGNE_GRIS = (230, 230, 230) # Gris très clair pour les lignes

# Mapping des comptes
NOMS_COMPTES = {
    1111: "Capital Social", 1481: "Emprunts bancaires", 2111: "Frais préliminaires",
    2340: "Matériel de transport", 2351: "Mobilier de bureau", 2352: "Matériel informatique",
    3111: "Marchandises (Stock)", 3421: "Clients", 3455: "État TVA Récupérable",
    4011: "Fournisseurs", 4411: "Fournisseurs", 4455: "État TVA Facturée",
    5141: "Banque (DH)", 5161: "Caisse", 6111: "Achats marchandises",
    6121: "Achats matières premières", 61251: "Achats Électricité", 61252: "Achats Eau",
    6131: "Location", 6142: "Transport", 6144: "Publicité", 6147: "Frais bancaires",
    6167: "Impôts et taxes", 6171: "Salaires", 7111: "Ventes marchandises",
    7121: "Ventes produits finis"
}

# --- CLASSE PDF "PRO" (Design Unifié) ---
class PDFPro(FPDF):
    def __init__(self, titre_doc):
        super().__init__()
        self.titre_doc = titre_doc

    def header(self):
        # Titre Principal (Style Journal)
        self.set_text_color(*COLOR_THEME) 
        self.set_font('Arial', 'B', 24)
        self.cell(100, 15, self.titre_doc, 0, 0, 'L')
        
        
        
        # Ligne de séparation Cyan épaisse
        self.ln(25)
        self.set_draw_color(*COLOR_THEME)
        self.set_line_width(0.8)
        self.line(10, 35, 200, 35)
        self.ln(10)

    def footer(self):
        # Pied de page style Journal (Bande Cyan)
        self.set_y(-15)
        self.set_fill_color(*COLOR_THEME)
        self.rect(0, 282, 210, 15, 'F') 
        self.set_font('Arial', 'I', 8)
        self.set_text_color(255, 255, 255)
        self.cell(0, 10, f'Page {self.page_no()}', 0, 0, 'C')

class GenerateurRapport:
    def _dossier(self, nom):
        path = os.path.join(BASE_EXP, nom.strip())
        if not os.path.exists(path): os.makedirs(path)
        return path

    def fmt(self, montant):
        """ Format Marocain : 10.000,00 """
        if montant is None: montant = 0.0
        txt = f"{montant:,.2f}"
        return txt.replace(",", "X").replace(".", ",").replace("X", ".")

    # --- EXPORT EXCEL (Inchangé) ---
    def exporter_journal_excel(self, data, nom):
        try:
            path = os.path.join(self._dossier(nom), f"Journal_{nom}.xlsx")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Journal"
            ws.append(["Date", "Libellé", "Compte", "Débit", "Crédit", "Référence"])
            
            fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
            font = Font(bold=True, color="FFFFFF")
            align = Alignment(horizontal="center", vertical="center")
            
            for c in ws[1]:
                c.fill = fill; c.font = font; c.alignment = align
            
            for row in data:
                l = list(row)
                l[3] = self.fmt(l[3]); l[4] = self.fmt(l[4])
                ws.append(l)
                for cell in ws[ws.max_row]: cell.alignment = align

            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['A'].width = 11
            ws.column_dimensions['C'].width = 8
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 16
            wb.save(path)
            wb.save(path)
            return path
        except: return None

    # --- NOUVEAU BILAN "STYLE JOURNAL" ---
    def generer_bilan_pdf(self, nom, l_act, t_act, l_pas, t_pas):
        try:
            date_jour = datetime.now().strftime("%d/%m/%Y")
            path = os.path.join(self._dossier(nom), f"Bilan_Style_{nom}.pdf")
            
            # Utilisation de la classe PRO
            pdf = PDFPro("BILAN")
            pdf.set_auto_page_break(auto=True, margin=20)
            
            # FONCTION INTERNE POUR DESSINER UNE PAGE
            def dessiner_page_bilan(titre_page, entetes, lignes, total_label, total_val):
                pdf.add_page()
                
                # 1. INFO SOCIETE (Style Journal)
                pdf.set_text_color(50, 50, 50)
                pdf.set_font('Arial', 'B', 14)
                pdf.cell(0, 10, f"Société : {nom.upper()}", 0, 1, 'L')
                
                pdf.set_font('Arial', 'B', 10)
                pdf.cell(15, 6, "Date : ", 0, 0)
                pdf.cell(50, 6, date_jour, 0, 1)
                pdf.ln(5)

                # 2. SOUS-TITRE (ACTIF ou PASSIF) CENTRÉ
                pdf.set_font('Arial', 'B', 16)
                pdf.set_text_color(*COLOR_THEME)
                pdf.cell(0, 10, titre_page, 0, 1, 'C')
                pdf.ln(2)

                # 3. EN-TÊTES TABLEAU (Fond Cyan, Texte Blanc)
                pdf.set_fill_color(*COLOR_THEME)
                pdf.set_text_color(*COLOR_TEXT_HEAD)
                pdf.set_font('Arial', 'B', 10)
                
                # Calcul largeurs (Style image bilan)
                # Actif: Libellé | Brut | Amort | Net
                # Passif: Libellé | Net
                if len(entetes) == 4:
                    w_lib, w_val = 80, 36 
                else:
                    w_lib, w_val = 140, 50

                pdf.cell(w_lib, 10, entetes[0], 0, 0, 'L', 1) # Align Left pour le libellé
                for h in entetes[1:]:
                    pdf.cell(w_val, 10, h, 0, 0, 'R', 1) # Align Right pour chiffres
                pdf.ln()

                # 4. LIGNES DU TABLEAU
                pdf.set_text_color(0, 0, 0)
                pdf.set_font("Arial", size=10)
                
                # Réinitialiser la couleur de ligne pour qu'elle soit fine et grise
                pdf.set_draw_color(*COLOR_LIGNE_GRIS)
                pdf.set_line_width(0.2)

                for i, l in enumerate(lignes):
                    # Zebra striping très léger
                    fill = True if i % 2 == 0 else False
                    pdf.set_fill_color(248, 248, 248) # Gris très très pâle

                    # Nom Compte
                    if l['compte'] == "RÉSULTAT NET": 
                        txt = "RÉSULTAT NET DE L'EXERCICE"
                        pdf.set_font("Arial", 'B', 10) # Résultat en gras
                    else:
                        n = NOMS_COMPTES.get(l['compte'], "")
                        txt = f"{l['compte']} - {n}" if n else str(l['compte'])
                        pdf.set_font("Arial", size=10)

                    # Affichage
                    pdf.cell(w_lib, 9, f"  {txt}", "B", 0, 'L', fill) # "B" = Bottom border only
                    
                    if len(entetes) == 4: # ACTIF
                        pdf.cell(w_val, 9, self.fmt(l['brut']), "B", 0, 'R', fill)
                        pdf.cell(w_val, 9, self.fmt(l['amort']), "B", 0, 'R', fill)
                        pdf.cell(w_val, 9, self.fmt(l['net']), "B", 0, 'R', fill)
                    else: # PASSIF
                        pdf.cell(w_val, 9, self.fmt(l['net']), "B", 0, 'R', fill)
                    
                    pdf.ln()

                # 5. TOTAL (Style Journal: Grand, Gras, Cyan)
                pdf.ln(5)
                pdf.set_font("Arial", 'B', 12)
                pdf.set_text_color(*COLOR_THEME)
                
                # Ligne épaisse au dessus du total
                y = pdf.get_y()
                pdf.set_draw_color(*COLOR_THEME)
                pdf.set_line_width(0.5)
                pdf.line(10, y, 200, y)
                pdf.ln(2)

                pdf.cell(w_lib, 10, total_label, 0, 0, 'L')
                
                if len(entetes) == 4:
                     pdf.cell(w_val, 10, self.fmt(total_val['brut']), 0, 0, 'R')
                     pdf.cell(w_val, 10, self.fmt(total_val['amort']), 0, 0, 'R')
                     pdf.cell(w_val, 10, self.fmt(total_val['net']), 0, 0, 'R')
                else:
                     pdf.cell(w_val, 10, self.fmt(total_val), 0, 0, 'R')
                
                # Double ligne de fin (Style comptable)
                pdf.ln(8)
                y = pdf.get_y()
                pdf.line(10, y, 200, y)
                pdf.line(10, y+1, 200, y+1)

            # --- GENERATION DES PAGES ---
            dessiner_page_bilan(
                "BILAN - ACTIF", 
                ["   RUBRIQUES / COMPTES", "BRUT", "AMORT/PROV", "NET"], 
                l_act, 
                "TOTAL GÉNÉRAL ACTIF", 
                t_act
            )

            dessiner_page_bilan(
                "BILAN - PASSIF", 
                ["   RUBRIQUES / COMPTES", "NET"], 
                l_pas, 
                "TOTAL GÉNÉRAL PASSIF", 
                t_pas
            )
            
            pdf.output(path)
            return path
        except Exception as e:
            print(f"Erreur PDF Bilan : {e}")
            return None

    # --- LIVRE JOURNAL (Déjà stylisé, on le garde) ---
    def generer_journal_pdf_style(self, nom_ent, data):
        try:
            date_jour = datetime.now().strftime("%d/%m/%Y")
            path = os.path.join(self._dossier(nom_ent), f"Livre_Journal_{nom_ent}.pdf")
            
            pdf = PDFPro("LIVRE-JOURNAL")
            pdf.set_auto_page_break(auto=True, margin=20)
            pdf.add_page()

            pdf.set_text_color(50, 50, 50)
            pdf.set_font('Arial', 'B', 14)
            pdf.cell(0, 10, f"Société : {nom_ent.upper()}", 0, 1, 'L')
            pdf.set_font('Arial', 'B', 10)
            pdf.cell(15, 8, "Date : ", 0, 0); pdf.cell(50, 8, date_jour, 0, 1); pdf.ln(5)

            pdf.set_fill_color(*COLOR_THEME); pdf.set_text_color(*COLOR_TEXT_HEAD); pdf.set_font('Arial', 'B', 10)
            pdf.cell(25, 8, "Date", 0, 0, 'L', 1)
            pdf.cell(90, 8, "Libellé / Compte", 0, 0, 'L', 1)
            pdf.cell(35, 8, "Débit", 0, 0, 'R', 1)
            pdf.cell(40, 8, "Crédit", 0, 1, 'R', 1)
            pdf.set_text_color(0, 0, 0)

            total_d, total_c = 0, 0
            dernier_grp = None 
            pdf.set_draw_color(*COLOR_LIGNE_GRIS)

            for row in data:
                date_op, libelle, compte_num, debit, credit, ref = row
                cle = f"{date_op}_{libelle}_{ref}"
                
                if cle != dernier_grp:
                    pdf.ln(3)
                    if dernier_grp:
                        y = pdf.get_y(); pdf.line(10, y, 200, y); pdf.ln(4)
                    pdf.set_font('Arial', 'B', 9)
                    pdf.cell(25, 6, str(date_op), 0, 0, 'L')
                    pdf.cell(100, 6, str(libelle), 0, 1, 'L')
                    dernier_grp = cle

                pdf.set_font('Arial', '', 9); pdf.set_text_color(80)
                nom = NOMS_COMPTES.get(compte_num, "")
                pdf.set_x(35)
                pdf.cell(15, 6, str(compte_num), 0, 0, 'L')
                pdf.cell(75, 6, nom, 0, 0, 'L')
                
                val_d = self.fmt(debit) if debit > 0 else ""
                val_c = self.fmt(credit) if credit > 0 else ""
                
                pdf.cell(30, 6, val_d, 0, 0, 'R')
                pdf.cell(35, 6, val_c, 0, 1, 'R')
                total_d += debit; total_c += credit

            pdf.ln(10)
            pdf.set_fill_color(*COLOR_THEME); pdf.set_text_color(255)
            pdf.set_font('Arial', 'B', 11)
            pdf.cell(115, 10, "TOTAUX", 0, 0, 'R', 1)
            pdf.cell(35, 10, self.fmt(total_d), 0, 0, 'R', 1)
            pdf.cell(40, 10, self.fmt(total_c), 0, 1, 'R', 1)

            pdf.output(path)
            return path
        except Exception as e:
            print(e); return None
